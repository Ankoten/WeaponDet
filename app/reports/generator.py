"""Генерация отчётов PDF и Excel."""
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

from app.config import OUTPUT_DIR
from app.storage.history import get_history, get_stats


def _get_cyrillic_font():
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        windir = os.environ.get("WINDIR", "C:\\Windows")
        font_candidates = [
            (os.path.join(windir, "Fonts", "arial.ttf"), "CyrillicFont"),
            (os.path.join(windir, "Fonts", "arial.TTF"), "CyrillicFont"),
            (os.path.join(windir, "Fonts", "Arial.ttf"), "CyrillicFont"),
            ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "CyrillicFont"),
            ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "CyrillicFont"),
        ]
        for font_path, font_name in font_candidates:
            if os.path.isfile(font_path):
                pdfmetrics.registerFont(TTFont(font_name, font_path, "UTF-8"))
                return font_name
        try:
            import reportlab
            rl_dir = Path(reportlab.__file__).resolve().parent
            vera = rl_dir / "fonts" / "Vera.ttf"
            if vera.exists():
                pdfmetrics.registerFont(TTFont("CyrillicFont", str(vera), "UTF-8"))
                return "CyrillicFont"
        except Exception:
            pass
        return None
    except Exception:
        return None


def generate_pdf(filepath: str | Path | None = None) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    path = Path(filepath) if filepath else OUTPUT_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)

    font_name = _get_cyrillic_font() or "Helvetica"

    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)
    styles = getSampleStyleSheet()
    for s in styles.byName.values():
        s.fontName = font_name

    story = []
    created_at = datetime.now().strftime("%d.%m.%Y %H:%M")

    title = "Report: Weapon Detection" if font_name == "Helvetica" else "Отчёт по детекции оружия"
    story.append(Paragraph(title, styles["Title"]))
    date_label = "Date:" if font_name == "Helvetica" else "Дата формирования:"
    story.append(Paragraph(f"{date_label} {created_at}", styles["Normal"]))

    stats = get_stats()
    history = get_history(limit=100)

    story.append(Spacer(1, 0.5 * cm))
    sect1 = "Statistics" if font_name == "Helvetica" else "Статистика"
    story.append(Paragraph(sect1, styles["Heading2"]))
    h1, h2 = ("Indicator", "Value") if font_name == "Helvetica" else ("Показатель", "Значение")
    stats_data = [
        [h1, h2],
        ["Total queries" if font_name == "Helvetica" else "Всего запросов", str(stats["total_queries"])],
        ["With weapon" if font_name == "Helvetica" else "С обнаружением оружия", str(stats["queries_with_weapon"])],
        ["Without weapon" if font_name == "Helvetica" else "Без обнаружения", str(stats["total_queries"] - stats["queries_with_weapon"])],
        ["Avg processing time (ms)" if font_name == "Helvetica" else "Среднее время обработки (мс)", str(stats["avg_processing_time_ms"])],
    ]
    for src, cnt in stats.get("by_source", {}).items():
        stats_data.append([f"By source: {src}" if font_name == "Helvetica" else f"По источнику: {src}", str(cnt)])
    t = Table(stats_data, colWidths=[9 * cm, 5 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#D9E2F3")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#D9E2F3"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)

    class_counts = Counter()
    for h in history:
        for d in h.get("detections", []):
            cls = d.get("class", "unknown")
            class_counts[cls] += 1
    if class_counts:
        story.append(Spacer(1, 0.8 * cm))
        sect2 = "Detected classes" if font_name == "Helvetica" else "Обнаруженные классы"
        story.append(Paragraph(sect2, styles["Heading2"]))
        ch1, ch2 = ("Class", "Count") if font_name == "Helvetica" else ("Класс", "Количество")
        class_rows = [[ch1, ch2]]
        for cls, cnt in class_counts.most_common():
            class_rows.append([cls, str(cnt)])
        t_class = Table(class_rows, colWidths=[6 * cm, 4 * cm])
        t_class.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#70AD47")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t_class)

    story.append(Spacer(1, 0.8 * cm))
    sect3 = "Recent queries (up to 50)" if font_name == "Helvetica" else "Последние запросы (до 50)"
    story.append(Paragraph(sect3, styles["Heading2"]))
    history = get_history(limit=50)
    if history:
        rh = ("Date", "Source", "File", "Detect.", "Time (ms)", "Classes") if font_name == "Helvetica" else ("Дата", "Источник", "Файл", "Детекц.", "Время (мс)", "Классы")
        rows = [list(rh)]
        for h in history:
            ts = h["timestamp"][:19].replace("T", " ") if h.get("timestamp") else "-"
            src = str(h.get("source", "-"))
            fn = (h.get("filename") or "-")[:25]
            dets = str(h.get("detections_count", 0))
            time_ms = str(round(h["processing_time_ms"], 1)) if h.get("processing_time_ms") is not None else "-"
            classes = ", ".join({d.get("class", "") for d in h.get("detections", [])})[:20] or "-"
            rows.append([ts, src, fn, dets, time_ms, classes])
        t2 = Table(rows, colWidths=[3.5 * cm, 2 * cm, 4 * cm, 1.5 * cm, 2 * cm, 3 * cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t2)
    else:
        empty_msg = "History is empty." if font_name == "Helvetica" else "История пуста."
        story.append(Paragraph(empty_msg, styles["Normal"]))

    doc.build(story)
    return path


def generate_excel(filepath: str | Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(filepath) if filepath else OUTPUT_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    stats = get_stats()
    history = get_history(limit=500)

    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Отчёт по детекции оружия"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells("A1:F1")

    row = 4
    ws[f"A{row}"] = "Статистика"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1
    for label, val in [
        ("Всего запросов", stats["total_queries"]),
        ("С обнаружением оружия", stats["queries_with_weapon"]),
        ("Без обнаружения", stats["total_queries"] - stats["queries_with_weapon"]),
        ("Среднее время обработки (мс)", stats["avg_processing_time_ms"]),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="По источникам")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for src, cnt in stats.get("by_source", {}).items():
        ws.cell(row=row, column=1, value=src)
        ws.cell(row=row, column=2, value=cnt)
        row += 1

    ws2 = wb.create_sheet("Классы детекций", 1)
    class_counts = Counter()
    for h in history:
        for d in h.get("detections", []):
            class_counts[d.get("class", "unknown")] += 1
    ws2["A1"] = "Класс"
    ws2["B1"] = "Количество"
    ws2["A1"].fill = header_fill
    ws2["A1"].font = header_font
    ws2["B1"].fill = header_fill
    ws2["B1"].font = header_font
    for r, (cls, cnt) in enumerate(class_counts.most_common(), start=2):
        ws2.cell(row=r, column=1, value=cls)
        ws2.cell(row=r, column=2, value=cnt)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14

    ws3 = wb.create_sheet("История", 2)
    headers = ["ID", "Дата/время", "Источник", "Файл", "Детекций", "Время (мс)", "Классы"]
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    for r, h in enumerate(history, start=2):
        classes = ", ".join({d.get("class", "") for d in h.get("detections", [])})
        ws3.cell(row=r, column=1, value=h.get("id", ""))
        ws3.cell(row=r, column=2, value=h.get("timestamp", ""))
        ws3.cell(row=r, column=3, value=h.get("source", ""))
        ws3.cell(row=r, column=4, value=(h.get("filename") or "")[:50])
        ws3.cell(row=r, column=5, value=h.get("detections_count", 0))
        ws3.cell(row=r, column=6, value=h.get("processing_time_ms"))
        ws3.cell(row=r, column=7, value=classes)
    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    ws4 = wb.create_sheet("Детали детекций", 3)
    det_headers = ["ID запроса", "Дата", "Источник", "Класс", "Уверенность %", "Кадр (сек)"]
    for col, h in enumerate(det_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    r = 2
    for h in history:
        qid = h.get("id", "")
        ts = (h.get("timestamp") or "")[:19]
        src = h.get("source", "")
        for d in h.get("detections", []):
            ws4.cell(row=r, column=1, value=qid)
            ws4.cell(row=r, column=2, value=ts)
            ws4.cell(row=r, column=3, value=src)
            ws4.cell(row=r, column=4, value=d.get("class", ""))
            ws4.cell(row=r, column=5, value=round(d.get("confidence", 0) * 100, 1))
            ws4.cell(row=r, column=6, value=d.get("time_sec", ""))
            r += 1
    for col in range(1, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    wb.save(str(path))
    return path
